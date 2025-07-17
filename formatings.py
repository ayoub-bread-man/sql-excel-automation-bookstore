
from openpyxl import load_workbook
from openpyxl.styles import Alignment , Font , PatternFill 
from the_log_p1 import write_log






def formating_logging( file_path  ) : 
    try : 

        wb = load_workbook( file_path)
        write_log("Report opened successfully")

    except : 
        write_log(" ERROR: Report file not found")
        print ("error the file doesn't exist ")
        exit()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]


        for cell in ws[1] :

            
            cell.alignment = Alignment(horizontal= 'center')
            cell.font = Font(bold=True, italic=True)
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold







        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # A, B, C...
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width



    wb.save(file_path)